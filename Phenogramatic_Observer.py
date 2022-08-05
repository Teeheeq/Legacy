# Begin by importing all modules that might be used in the script


#Just a library of preferred imports to cover my bases
#Importing os module
import os
#shutil is another library module that carries out a similar function
import shutil
#importing date module
#importing time module
from datetime import time
import time
#importing random module
import random
from webbrowser import get
#Excel handling module
import xlrd
import xlwt
import pandas
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import openpyxl

#Setting up the directories that will be acccessed throughout the script


#Setup a standard home directory for Excel / csv manipulation
home_dir = os.path.join("T:\\", "Project Blue Rose", "Git Repo", "Legacy", "Pioneer")
#set the filename for the pre-formatting excel file
pre_excel_name = "Con_1"
post_excel_name = "Output"
#amend the home directory for access to the excel file
pre_excel_path = os.path.join(home_dir, pre_excel_name + ".xlsx")
source_csv_path = os.path.join(home_dir, pre_excel_name + ".csv")



#Defining specific global shells
formatted_numbers = []
list_of_phone_numbers_from_excel = []
formatted_postcodes = []
list_of_postcodes_from_excel = []
formatted_names = []
list_of_names_from_excel = []
formatted_addresses = []
list_of_adresses_from_excel = []


#This is only called after files are verified to exist, to prevent huge data errors
#this function is the master call funtion, this ensures all other functions are called in the correct order and before the file is saved.
def function_caller():
    #things that need to be formatted first please!
    name_fetcher()
    phone_number_fetcher()
    postcode_fetcher()
    address_fetcher()
    
    #stuff that we have to generate ourselves
    country_injector()
    weight_injector()
    parcel_no_injector()
    SKU_injector()
    
    
    worksheet.delete_cols(1)
    workbook_openpyxl.save("Output.xlsx")
    
    #saving our amazing work
    panda_workbook_access = pandas.read_excel("Output.xlsx")
    panda_workbook_access.to_csv("Output.csv",  index=False)
    
    os.remove(pre_excel_path)
    os.remove(source_csv_path)
    os.remove("Output.xlsx")
    
    print("All done!")
    #printing a message to the user to show that the script has finished
    
    






#This function will create files if needed
def xlsx_generation(new_excel_file_needed, source_csv_exists):
    
    #There are 4 possible scenarios for the excel file and CSV file status
    #This function will address and inform the End_User of the action_paths that will be taken
    
    
    #Case 1 of 4: The excel file does not exist, but the csv file does
    #If the excel file does not exist, but we do have a source, create one to be used
    if new_excel_file_needed and source_csv_exists:
        #if the excel file does not exist and the csv file does, the csv file is copied to the excel file
        read_csv = pandas.read_csv(os.path.join(home_dir, pre_excel_name + ".csv"))
        read_csv.to_excel(pre_excel_path)
        print("""
              Excel conversion successful:
              The Excel file was successfully converted from the csv file
              the file is located at:
              """
              + pre_excel_path)
        
    #Case 2 of 4: The excel file does not exist, and the csv file does not exist    
    #We need a new Excel file, but we dont have anything to generate it from, unfortunately this is only solved by the End_User
    elif new_excel_file_needed and not source_csv_exists:
        print("""
              The Excel file does not exist, but neither does a source CSV file
              Unfortunately I am unable to generate pending dispaches on my own, please place a source CSV at""" 
              + source_csv_path + """ and try again""")
        print("The program will now be terminated")
        time.sleep(5)
        os._exit(0)
        
    #Case 3 of 4: The excel file does exist, but the csv file does not...
    #CSV file exists!
    #...
    #but so does the excel file...
    #that... isn't intended....
    #We have no clue why this would ever happen, so we just inform the End_User, and provide them with some information that may be of use to them
    elif not new_excel_file_needed and source_csv_exists:
        print("""The source CSV file had been found!""")
        time.sleep(1)
        print("""But it seems that the Excel file also already exists...""")
        time.sleep(2)
        print("""Strange that... That shouldn't really happen...""")
        time.sleep(3)
        print("Let me find some more information that may be of use to you:")
        time.sleep(2)
        csv_creation_time = os.path.getctime(source_csv_path)
        excel_creation_time = os.path.getctime(pre_excel_path)
        if csv_creation_time > excel_creation_time:
            print("""
                  The CSV file was created more recently than the Excel file
                  This means that the Excel file is outdated, and will be overwritten
                  Please make sure that you have a backup of the Excel file before proceeding
                  """)
            time.sleep(2)
        elif csv_creation_time < excel_creation_time:
            print("""
                  The Excel file was created more recently than the CSV file
                  This means that the CSV file is likely to be it's source,
                  We can proceed as normal
                  """)
            time.sleep(2)
        
    # skipping case 4 for time purposes, due to expediting,
    # if returning to bolster, worth rechecking

        
#defining a function that will check whether a Con_1 csv and/or excel file exists
def check_for_existing_files():
    #checking for the existence of the excel file
    if os.path.exists(pre_excel_path):
        print("Excel file found")
        #First definition will be False, as the excel file exists
        new_excel_file_needed = False
    else:
        print("Excel file not found")
        #First definition will be True, as the excel file needs to be created in the current directory
        new_excel_file_needed = True
        
    #checking for the existence of the csv file
    if os.path.exists(os.path.join(home_dir, pre_excel_name + ".csv")):
        print("CSV file found")
        #First definition will be True as the csv file exists in the current directory
        source_csv_exists = True
    else:
        print("CSV file not found")
        #First definition will be False as the csv file does not exist in the current directory
        source_csv_exists = False
    xlsx_generation(new_excel_file_needed, source_csv_exists)








#creates data for the "Buyer Country" column
def country_injector():
    #Cycling through the rows in the "Buyer Country" of the excel file
    
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "L" + str(i+1)] = "United Kingdom"
    print("Added Buyer Country") 
    
    
#creates data for the "No of Parcels" column   
def parcel_no_injector():
    #Cycling through the rows in the "No of Parcels" of the excel file
    
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "N" + str(i+1)] = "1"
    print("Added No of Parcels")     
    
#creates data for the "Weight" column 
def weight_injector():
    #Cycling through the rows in the "Weight" of the excel file
    
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "O" + str(i+1)] = "25"
    print("Added Weight") 
    
    
#creates data for the "SKU" column   
def SKU_injector():
    #Cycling through the rows in the "SKU" of the excel file
    worksheet["P1"] = "SKU"
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "P" + str(i+1)] = "FX5"
    print("Added SKU")     
        






        
#function that will write back and replace the existing phone numbers with the correctly formatted ones
def phone_number_injector(formatted_phone_numbers):
    #Cycling through the rows in the phone number column of the excel file
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ get_column_letter(5) + str(i+1)] = formatted_phone_numbers[i-1]
    print("Phone numbers successfully formatted")

#funtion used to format phone numbers
def phone_number_format(phone_numbers):
    #format the phone number to be in the format of xxxx xxx xxx (without spaces), where x denotes an integer between 0-9 inclusive
    banned_character_for_phone = ["-", " ", "(", ")", "+", ".", "/", "*", "#", "!", "@", "$", "%", "^", "&", "?", "!", '"',	"#", "$", "%", "&",	"'", "(", ")", "*", "+", ",","-",".", "/", ":", ";"	,"<", "=", ">", "?", "~", "]", "["]
    
    #checking each number individually
    for number in phone_numbers:
        #first dealing with weird excel formatting
        if number != None and ("E+11" or "e+11") in number:
            number = number.replace("E+11", "")
            number = "".join(c for c in number if c.isdecimal())
            number = str(int(float(number) * (10**11)))
        
        
        #removing special characters from the phone numbers
        number = "".join(c for c in number if c.isdecimal())  
        #If there is a null feild, replace it with valid 0 string
        if number == "0":
            number = "1582290000"
        if number == "":
            number = "1582290000"
        
        
        
        #exploding into an array of characters so that we can check each character individually
        exploded_number = list(number)
        #so we dont create a None respone
        if(len(exploded_number) > 3):
            #first checkong for "0" or "00" at the start of the number
            for i in exploded_number[0:2]:
                
                if i == "0":
                    exploded_number.remove(i)
            #dealing with "+44"
            if str(exploded_number[0] + exploded_number[1]) == "44":
                #remove the first two digits of the number
                exploded_number.remove(exploded_number[0])
                #old number 1 is now 0 now that old 0 is removed  
                exploded_number.remove(exploded_number[0])
       
        number = "".join(exploded_number)
        if len(number) != 10:
            length = len(number)
            if length > 10:
                number = number[0:10]
            elif length == 0:
                for i in range(0,10):
                    exploded_number.append(random.randint(0,9))
            else:
                for i in range(0,10-length):
                    exploded_number.append("0")
                number = "".join(exploded_number)
        
        formatted_numbers.append(number)
    phone_number_injector(formatted_numbers)

     
#This function fetches the phone numbers from excel and places them into an array to be sent for formatting purposes
def phone_number_fetcher():
    #Cycling through the rows in the phone number column of the excel file
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, a new entry is created in the list_of_phone_numbers_from_excel array
        if sheet.cell_value(i, 4) == "":
            list_of_phone_numbers_from_excel.append("0")
        if sheet.cell_value(i,4) != "":
            list_of_phone_numbers_from_excel.append(str(sheet.cell_value(i, 4)))
        
    #using the list to call for a formatted output
    phone_number_format(list_of_phone_numbers_from_excel)
    
   
   
   
   
   
   
#recieves the formatted_postcodes and injects them into the excel file 
def postcode_injector(formatted_postcodes):
    #Cycling through the rows in the "Buyer Posdtcode" of the excel file
    for i in range(1 , sheet.nrows -1):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "K" + str(i+1)] = formatted_postcodes[i-1]
    print("Postcodes successfully formatted")

  
  
#splices the fetched postcode, sends them for injection into the excel file
def postcode_splicer(postcodes):
    for postcode in postcodes:
        exploded_postcode = (list(postcode))
        length = len(exploded_postcode)
        reversed_exploded_postcode = list(postcode[::-1])
        if length != 0:
            if reversed_exploded_postcode[3] == " ":
                formatted_postcodes.append("".join(exploded_postcode))
            elif reversed_exploded_postcode[3] != " ":
                stringed_explosion = "".join(reversed_exploded_postcode)
                stringed_explosion = stringed_explosion[:3] + " " + stringed_explosion[3:]
                formatted_postcodes.append("".join(stringed_explosion[::-1]))
    postcode_injector(formatted_postcodes)
            

#Fetches postcodes for splicing and verification
def postcode_fetcher():
    #Cycling through the rows in the phone number column of the excel file
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, a new entry is created in the list_of_postcodes_from_excel array
        list_of_postcodes_from_excel.append(str(sheet.cell_value(i,10)))
    
    postcode_splicer(list_of_postcodes_from_excel)
    



#receives the formatted names and injects them into the excel file
def name_injector(formatted_names):
    #Cycling through the rows in the "Buyer Name" of the excel file
    for i in range(1 , sheet.nrows -1):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "D" + str(i+1)] = formatted_names[i-1]
    print("Names successfully formatted")


#receives the names and formats them for injection
def name_formatter(names):
    for name in names:
        #removing special characters from the names
        name = "".join(c for c in name if c.isalpha() or c==" ")
        formatted_names.append(name)
    name_injector(formatted_names)
    

#Fetches names and appends them to a list for formatting purposes
def name_fetcher():
    #Cycling through the rows in the "Buyer Name" column of the excel file
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, a new entry is created in the list_of_postcodes_from_excel array
        list_of_names_from_excel.append(str(sheet.cell_value(i,3)))
    name_formatter(list_of_names_from_excel)
    
    
 
#receives the formatted addresses and injects them into the excel file
def address_injector(formatted_addresses):
     #Cycling through the rows in the "Buyer Address" of the excel file
    for i in range(1 , sheet.nrows -1):
        #for each "row"/cell in that column, an entry is replaced with it's formatted value
        worksheet[ "G" + str(i+1)] = formatted_addresses[i-1]
    print("Addresses successfully formatted")
 
 
    
#reveives the list of addresses and formats them for injection
def address_formatter(addresses):
    for address in addresses:
        #removing special characters from the addresses
        exploded_address = list(address)
        for i in exploded_address:
            if i == "-":
                exploded_address[exploded_address.index(i)] == " "
        address = "".join(c for c in address if c.isalpha() or c==" " or c.isdecimal() or c=="-")
        address.replace("-", " ")
        address_shortened = address[:34]
        formatted_addresses.append(address_shortened)
    address_injector(formatted_addresses)
    
    
#Fetches the adresses and appends them to a list for formatting purposes    
def address_fetcher():
    #Cycling through the rows in the "Buyer Address 1" column of the excel file
    for i in range(1 , sheet.nrows):
        #for each "row"/cell in that column, a new entry is created in the list_of_adresses_from_excel array
        list_of_adresses_from_excel.append(str(sheet.cell_value(i,6)))
    address_formatter(list_of_adresses_from_excel)
    

check_for_existing_files()
#open the workbook for general callback
workbook = xlrd.open_workbook(pre_excel_path)
workbook_openpyxl = openpyxl.load_workbook(pre_excel_path)
workbook_openpyxl.sheetnames
worksheet = workbook_openpyxl["Sheet1"]
workbook_write = xlwt
sheet = workbook.sheet_by_index(0)
function_caller()

