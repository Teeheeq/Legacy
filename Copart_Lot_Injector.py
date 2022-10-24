#================================================================================================================================================================================================================================================================================================================================================================================
#
#Version: TQMK_0.0.1
#Author: TQ
#Date: 16-09-2022
#Project: Ford-Compiler
#Project Goals:
#   Pulls data from Copart
#   Appends them to the spreadsheet using minimal space
#   Checks for errors and duplicates
#   Should be able to run multiple times without errors
#   Independant of End_User
#   Should be structured to be easily run on a daily basis as well as being Modular
#   Should be able to be run on a server
#   Will use PyPDF as a supplicant to the main program
#
#Patch Notes:
#   -Began Project on 16-09-2022
#   -Licensed Software to Ford Spare Parts Ltd
#       Unit 16
#       Station Approach
#       Hitchin
#       Hertfordshire
#       SG4 9UW
#       Tel: 01462 832 834
#
#================================================================================================================================================================================================================================================================================================================================================================================
#
#   Legal Disclaimer, License and Terms of Use
#
#================================================================================================================================================================================================================================================================================================================================================================================
#
#   This program is licensed to Ford Spare Parts Ltd; This program and any attachment(s) are intended for the above named only and should not be distributed to any other party without the express permission of the author.
#   This program is not to be used for any other purpose than that for which it was intended. If you are the recipient of this program and you are not Ford Spare Parts Ltd, you are not permitted to use this program.
#   Ford Spare Parts Ltd are not permitted to distribute this program to any other party without the express permission of the author.
#   Any breach of this agreement will result in the immediate termination of the license and the author will be entitled to pursue any legal action against the breaching party.
#   The author will not be held responsible for any loss or damage caused by the use of this program.
#   Please note that to ensure regulatory compliance, this program is subject to change at any time without notice; The supplimental ID file is subject to complete secrecy and any third party who gains access to this file will be prosecuted to the full extent of the law.
#   It is the responsibility of the End User to ensure that the program is used in accordance with the terms and conditions of the license agreement.
#   Issue of this program does not constitute a license to use any intellectual property rights owned by the author, nor does it constitute a license to use any intellectual property rights owned by any third party.
#   This program is provided "as is" without warranty of any kind, either express or implied, including, but not limited to, the implied warranties of merchantability, fitness for a particular purpose, or non-infringement.
#   The author does not warrant or make any representations regarding the use or the results of the use of the program in terms of correctness, accuracy, reliability, or otherwise.
#   The End User assumes the entire risk as to the quality and performance of the program.
#   The author does not warrant that the functions contained in the program will meet the End User's requirements or that the operation of the program will be uninterrupted or error-free.
#   The End User must pass this agreement on when distributing the program, within the site license.
#   The End User must not remove or alter any copyright notices on any and all parts of the program; The End User must not remove or alter any trademark notices on any and all parts of the program.
#   The End User must not remove or alter any license notices on any and all parts of the program.
#   The End User must not remove or alter any warranty notices on any and all parts of the program.
#   The End User must not remove or alter any disclaimer notices on any and all parts of the program.
#   The End User must not remove or alter any limitation of liability notices on any and all parts of the program.
#   The End User must not remove or alter any indemnity notices on any and all parts of the program.
#   The End User must not remove or alter any other notices on any and all parts of the program.
#   Should a third party claim that the program or any part thereof infringes that third party's intellectual property rights, the End User will be responsible for defending such claim.
#   The End User must indemnify the author against any damages, costs, and expenses, including reasonable attorneys' fees, incurred by the author as a result of any claim or action brought against the author by a third party alleging that the program or any part thereof infringes that third party's intellectual property rights.
#   If the End User intends to take on another developer to work on the program, the End User must inform the author of this fact and the author must agree to this fact in writing.
#   The End User must not use the program in any way that is unlawful, illegal, fraudulent or harmful, or in connection with any unlawful, illegal, fraudulent or harmful purpose or activity.
#   If the End User fails to comply with any provision of this agreement, the End User's license to use the program will automatically terminate.
#   Usage or possesion of this program is subject to the terms and conditions of the license agreement; By using this program, you are agreeing to the terms and conditions of the license agreement, as stated above.
#   Refusal to comply with the Terms and Conditions of the license agreement will result in the immediate termination of the license and the author will be entitled to pursue any legal action against the breaching party, as well as any other party who has been granted access to this program.
#   
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


#================================================================================================================================================================================================================================================================================================================================================================================
#
#   IMPORTING MODULES
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from enum import auto
from http.client import TOO_MANY_REQUESTS
from multiprocessing.sharedctypes import Value                      #   For handling 429 errors with the Copart website
import os
from queue import Empty                                             #   Neccesary to handle system functions
import shutil                                                       #   For transferring files from the Downloads folder to the Payments_History folder
import datetime                                                     #   Makes isinstance(date, datetime.date) much easier to handle
from datetime import date, timedelta
from sqlite3 import DatabaseError                                   #   Important to parse 'Payments_History.xlsx' from other downloads and days
import numpy as np                                                  #   Used in np.nan calls in order to handle the DataFrame cleanly 
import time                                                         #   Necessary for sleep calls to work
import xlrd                                                         #   Function to convert the downloaded CSV file to an XLSX file, as well as reading the XLSX file
import pandas                                                       #   The backbone of DataFrame and Excel manipulation
from selenium import webdriver                                      #   Webdriver Functions
from selenium.webdriver.common.by import By                         #   Enables Element Selection by various methods
from selenium.webdriver.support.ui import WebDriverWait             #   Enables the script to wait for the page to load before continuing
from selenium.webdriver.support import expected_conditions as EC    #   Hands WebDriverWait the element to wait for, and the time to wait for it to appear
from selenium.webdriver.support.ui import Select                    #   Allows webdriver to parse and manipulate list and table elements
from selenium.webdriver.common.action_chains import ActionChains    #   ActionChains give the script a less elegant way of performing M&K actions, should automation fail
from selenium.common.exceptions import TimeoutException             #   Helps the script determine Exceptions and handle them accoringly; Frequently associated with webdriver
from openpyxl import load_workbook                                  #   Enables pandas to read the 'Payments_History.xlsx' file and generate a DataFrame
from openpyxl.utils.cell import get_column_letter                   #   Enables readback to 'COD_Vehicle_List.xlsx' without needing to manually list headers
import openpyxl                                                     #   Neccesary for appending script DataFrame to 'COD_Vehicle_List.xlsx'
import PyPDF2                                                       #   Parses and reads through PDFs in 'Payments_History' folder; Fetches 'Total Reg' and 'CAT' values
import json                                                         #   Helps the web.ExecuteScript() function execute the Javascript code within webdriver
import msvcrt                                                       #   Needed to create Automation
import pyfiglet                                                     #   Used in Override Hub for Welcome
import sys                                                          #   Preferred kill mechanism for the script; Prevents the script from running if the user asks it to end.
from pandas.api.types import CategoricalDtype                       #   Used for custom sort
import re                                                           #   Used to ensure regular format
#================================================================================================================================================================================================================================================================================================================================================================================


#================================================================================================================================================================================================================================================================================================================================================================================
#
#   DEFINING GLOBALS
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#   webdriver options; used to ensure window.print() calls "Save As" instead of "Print"
#
chrome_options = webdriver.ChromeOptions()                          #   This was yoinked from a kind StackOverflow post; I personally hadn't thought of reassigning default print settings, so kudos to the author!
settings = {
       "recentDestinations": [{
            "id": "Save as PDF",
            "origin": "local",
            "account": "",
        }],
        "selectedDestinationId": "Save as PDF",
        "version": 2
    }
prefs = {'printing.print_preview_sticky_settings.appState': json.dumps(settings)}
chrome_options.add_experimental_option('prefs', prefs)
chrome_options.add_argument('--kiosk-printing')
chrome_options.add_argument("--log-level=3")
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#   Assigning Directory Paths and fetching keys to be passed to the webdriver in Copart_login()
#
curr_dir = os.getcwd()                                              #   Fetches current directory from X_id.py; If the line reads as 'curr_dir = os.getcwd()', the script is running on TQ's laptop as it is being debugged, if so please rectify.
os.chdir(curr_dir)                                                  #   Sets Relative Run Path for the script
downloads_dir = os.path.expanduser("~")+ str("\Downloads")          #   Fetches User Paths from OS
import FSP_XID
username = FSP_XID.username                                         #   Fetches username from X_id.py
password = FSP_XID.password                                         #   Fetches password from X_id.py
#
#================================================================================================================================================================================================================================================================================================================================================================================


#================================================================================================================================================================================================================================================================================================================================================================================
#
#   FUNCTION DESCRIPTOR DEFINITIONS
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#   <Preposition> Module: <Name>
#           - This script is broken down into <Preposition> Module Groups, each of which is collectively responsible for a specific task or block of tasks for each stage of execution.
#           - The <Preposition> tags are in reference to the Module's position in the script, and are used to help the Developer understand the order of execution.
#
#   Purpose: <Purpose>
#           - <Purpose> denotes the intended function of the Module, and what should be achieved by the end of it's run.
#
#   Elevated Caller: <Function>
#           - <Function> denotes the function that has a larger an impact on the Module's intended function.
#           - This function will string together several function calls in an attempt to streamline the Module's execution.
#           - It is possible that there are more than one Elevated Callers working in tandem within a Module.
#
#   Functions: <Functions>
#           - <Functions> denotes the functions that are defined and used witihin a Module.
#           - It is important to note that this will NOT include any Elevated Caller functions.
#
#   Parameters: <Parameters>
#           - <Parameters> denotes the parameters that are passed to a function, and are necessary for the function to run.
#
#   Order of Execution: <Order of Execution>
#           - <Order of Execution> denotes the order in which the Module's functions are to be executed.
#           - If any functions are completely nested within other functions, this is displayed as a nested list.
#               - This means that if a nested function is called within several functions, it will be seen more in more than one nest.
#           - If a function *may* be called from a module as a selection from a list, the list will be preceeded by a "-->" indent.
#
#   Connection Call: <Function> Calls <Function>
#           - <Function> Calls <Function> denotes the function in the Module that calls a function from the succeeding Module, thereby being the Connection call.
#
#================================================================================================================================================================================================================================================================================================================================================================================


#================================================================================================================================================================================================================================================================================================================================================================================
#
#   FUNCTIONS
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#                                               Precursor Module: 90 Day CSV Fetcher    
#
#           Purpose: To fetch the CSV file from Copart containing the last 90 days of Purchases
#
#           Elevated Caller: 
#               copart_90_day_csv_fetcher()
#                               -Parameters: None
#
#           Functions:
#               copart_login()              
#                       - Logs into Copart and navigates to the 'Payments History' page
#                               -Parameters: None
#               nav_to_payment_page()
#                       - Navigates to the 'Payments History' page
#                               -Parameters: web
#               csv_downloader()
#                       - Downloads the CSV file from the 'Payments History' page
#                               -Parameters: web
#               csv_to_xlsx()
#                       - Converts the downloaded CSV file to an XLSX file and places it in the current directory
#                               -Parameters: None
#
#           Order of Execution:
#               copart_90_day_csv_fetcher()
#               copart_login()
#               nav_to_payment_page()
#               csv_downloader()
#               csv_to_xlsx()
#               
#           Connection call:
#               copart_90_day_csv_fetcher()
#                   - Calls the generate_dataframe() function, located in the 'Dataframe Generator' Module
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function to login to Copart;
#   Utilizes webdriver, By, WebDriverWait, ExpectedConditions and time
#
#
def copart_login():
    global web                                                                                              #   Setup of webdriver has been made global to allow it to be used in other functions, regardless of a parameter call
    web = webdriver.Chrome(chrome_options=chrome_options)                                                   #   webdriver is set to use the chrome_options defined above
    web.get("https://Copart.co.uk/login")                                                                   #   Opens the Copart website to the login page
    web.maximize_window()                                                                                   #   Maximises the window to ensure the script is able to run without a ClickIntercept error
    WebDriverWait(web, 100).until(EC.presence_of_element_located((By.ID, "username")))                      #   Ensures the username field is located; Waits for 100 seconds for the field to be located before throwing a TimeoutException error
    time.sleep(0.5)                                                                                         #   Sleeps for 0.5 seconds to allow the webpage to load, just in case
    web.find_element(By.ID, "username").send_keys(username)                                                 #   Sends the username to the username field
    WebDriverWait(web, 100).until(EC.presence_of_element_located((By.ID, "password")))                      #   Ensures the password field is located; Waits for 100 seconds for the field to be located before throwing a TimeoutException error
    time.sleep(0.5)                                                                                         #   Sleeps for 0.5 seconds to allow the webpage to load, better safe than sorry, right?
    web.find_element(By.ID, "password").send_keys(password)                                                 #   Sends the password to the password field
    web.find_element(By.CLASS_NAME, "loginfloatright").click()                                              #   Clicks the login button
    time.sleep(5)                                                                                           #   Gives the webpage a good 5 seconds to load the redirect to "https://Copart.co.uk/dashboard"
    try:                                                                                                    #   If the "Accept Cookies" button is present, click it; This avoids a potential later ClickIntercept error
        time.sleep(3)                                                                                       #   Wait to be sure the webpage has loaded
        web.find_element(By.ID, "onetrust-accept-btn-handler").click()                                      #   Clicks the "Accept Cookies" button
    except Exception:                                                                                       #   If the "Accept Cookies" button is not present, continue
        pass
    return(web)                                                                                             #   Returns the webdriver to the calling function; Yes, I know it was assigned globally, just a habit...
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for navigating to the 'Payments History' page; Sets the Date range and Number of items to 90 and 100 respecively;
#   Utilizes webdriver, By, WebDriverWait, ExpectedConditions and time
#
#
def nav_to_payment_page(web):
    try:                                                                                                    #   Maximises the window to ensure the script is able to run without a ClickIntercept error
        web.maximize_window()
    except Exception:                                                                                       #   If the window is already maximised, continue
        pass
    try:                                                                                                    #   If the "Accept Cookies" button is present, click it; This avoids a potential later ClickIntercept error
        web.find_element(By.ID, "onetrust-accept-btn-handler").click()                                      #   'Cicks it'
    except Exception:                                                                                       #   If the "Accept Cookies" button is not present, continue
        pass                                                                                                #   The below lines are used to set the Date range and Number of items to 90 and 100 respectively
    web.get("https://www.copart.co.uk/member-payments/payment-history?viewType=invoice")                    #   Navigates to the 'Payments History' page; This is our script's homepage of sorts; We'll be coming back here often as this page has all of the data we need
    WebDriverWait(web, 100).until(EC.presence_of_element_located((By.CLASS_NAME, "payment-filters-dropdown")))
    web.find_element(By.CLASS_NAME, "payment-filters-dropdown").click()                                     #   Clicks the Payment History Date dropdown; I.E 'Last 30 Days', 'Last 60 Days', 'Last 90 Days', etc etc
    time.sleep(1)                                                                                           #   Waits for the dropdown to be visible
    web.execute_script("arguments[0].scrollIntoView();", web.find_element(By.XPATH, "//*[contains(text(), 'Last 90 Days')]"))
    web.find_element(By.XPATH, "//*[contains(text(), 'Last 90 Days')]").click()                             #   Clicks the 'Last 90 Days' option; This is because the next option is 90 to 180 days, we want recent data, so we'll have to make do with 90 days.
    web.execute_script("arguments[0].scrollIntoView();", web.find_element(By.CLASS_NAME, "p-paginator-rpp-options"))
    web.execute_script("arguments[0].click();", web.find_element(By.CLASS_NAME, "p-paginator-rpp-options")) #   Clicks the 'Show dropdown' button to select 100 items
    time.sleep(1)                                                                                           #   Waits for the dropdown to be visible
    web.execute_script("arguments[0].scrollIntoView();", web.find_element(By.CLASS_NAME, "p-dropdown-item"))#   Scrolls to the top of the options list; We are trying to avoid a ClickIntercept error
    time.sleep(1)                                                                                           #   Gives the webpage a good second to load the dropdown options
    web.find_element(By.XPATH, "//li[@aria-label='100']").click()                                           #   Sets the Number of items to 100
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for downloading the CSV, 'Payment History' file from the 'Payments History' page;
#   Utilizes webdriver, By, WebDriverWait, ExpectedConditions and time   
#
#   
def csv_downloader(web):
    WebDriverWait(web, 100).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Export')]")))
    time.sleep(5)                                                                                           #   Waits for the Export button to be visible
    web.find_element(By.XPATH, "//*[contains(text(), 'Export')]").click()                                   #   Clicks the Export button
    time.sleep(5)                                                                                           #   Waits for the CSV to download before any other functions reference it; This avoids a FileNotFound error
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for converting the downloaded CSV file to an XLSX file and placing it in the current directory;
#   Utilizes date, time, shutil, pandas and os
#
#
def csv_to_xlsx():
    filename_date = date.today().strftime("%Y %B %d")                                                       #   Sets the filename_date to the current date in the format 'YYYY Month DD', used in search for the CSV file
    time.sleep(3)
    try:                                                                                                    #   If the CSV file is present, convert it to an XLSX file and place it in the current directory
        shutil.move(downloads_dir+"\PaymentsHistory_Invoice_"+filename_date+ ".csv", curr_dir+"\PaymentsHistory.csv")
        curr_csv = curr_dir+"\PaymentsHistory.csv"
        pandas.read_csv(curr_csv).to_excel(curr_dir+"\PaymentsHistory.xlsx", index=False)
        os.remove(curr_csv)
    except Exception:
        pass                                                                                                #   If the CSV file still exists in the downloads directory, delete it
    if os.path.exists(downloads_dir+"\PaymentsHistory_Invoice_"+filename_date+ ".csv"):
        os.remove(downloads_dir+"\PaymentsHistory_Invoice_"+filename_date+ ".csv")
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Elevated Caller Function; Calls the other functions in the Precursor Module Group
#   Utilizes all Daughter imports; Refer to the Daughter Module for more information
#
#
def copart_90_day_csv_download():
    copart_login()                                                                                          #   Logs into Copart
    nav_to_payment_page(web)                                                                                #   Navigates to the 'Payments History' page
    csv_downloader(web)                                                                                     #   Downloads the up-to-date Payment History
    csv_to_xlsx()                                                                                           #   Converts the downloaded CSV file to an XLSX file and places it in the current directory
    generate_dataframe()                                                                                    #   Calls the generate_dataframe function to generate the dataframe from the XLSX file
#                                       
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------  
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#                                               Contiguous Module: Dataframe Generator
#
#           Purpose: Generates a DataFrame of Vehicles with Undocumented Lot/Inv_# Numbers; Injects them into the 'COD_Vehicle_List.xlsx' file
#
#           Elevated Caller:
#               df_caller()
#                               -Parameters: df
#           Functions:
#               generate_dataframe()
#                       - Creates an empty DataFrame; This will be flushed with the data from the XLSX file
#                               -Parameters: None
#               df_append()
#                       - Appends the data from the XLSX file to the empty DataFrame
#                               -Parameters: df, copart_header, target_header
#               array_generator()
#                       - Creates an array of the data from the XLSX file and Parameters given
#                               -Parameters: array_header
#               df_formatter()
#                       - Formats the data in the DataFrame to be standardised for later parsing
#                               -Parameters: df
#               xlsx_injector()
#                       - Prepares the DataFrame for injection into the 'COD_Vehicle_List.xlsx' file
#                               -Parameters: df
#
#           Order of Execution:
#               generate_dataframe()
#               df_caller()
#               |   df_append()
#                   |   array_generator()
#               df_formatter()
#
#           Connection Call:
#               xlsx_injector()
#                   - Calls the Invoice_fetcher() function, located in the 'Invoice Fetcher' Module
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for generating the DataFrame;
#   Utilizes pandas
#
#
def generate_dataframe():                                                                                   #   Creates an empty DataFrame; This has the columns needed in order
    df = pandas.DataFrame({
        "Date": [],
        "Lot/Inv_#": [],
        "Registration": [],
        "VIN": [],
        "Make": [],
        "Vehicle": [],
        "Colour": [],
        "Location": [],
        "Category": [],
        "LLT": []
    })
    df_caller(df)
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function to call for generation of DataFrame arrays, and their injection into the DataFrame;
#   Does not contain any imports
#
#
def df_caller(df):                                                                                          #   Calls the df_append function for each column needed in the DataFrame
    df_append(df, "Invoice Date", "Date")
    df_append(df, "Model", "Vehicle")
    df_append(df, "Make", "Make")
    df_append(df, "Lot/Inv #", "Lot/Inv_#")
    df_append(df, "VIN", "VIN")
    df_append(df, "Location", "Location")
    df_formatter(df)
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for appending the data from the XLSX file to the requested column in the DataFrame;
#   Utilizes pandas
#
#
def df_append(df, copart_header, target_header):    
    df[target_header] = array_generator(copart_header)                                                      #   Using the parameters provided, asks array_generator for an array of the requested data, and then appends it to the DataFrame in the Column requested
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for generating the array of the data from the XLSX file using the parameter given;
#   Utilizes xlrd
#
#
def array_generator(array_header, sheet = xlrd.open_workbook(curr_dir+"\PaymentsHistory.xlsx").sheet_by_index(0), ascending = False):
    requested_array = []                                                                                    #   Creates an empty array, which will be filled with the data from the XLSX file
    relevant_column = None                                                                                  #   We havent yet got a column to work with, so set it to None
    if ascending == False:
        step = -1                                                                                           #   If the data is to be sorted in descending order, set the step to -1
        start = sheet.nrows -1                                                                              #   If the data is to be sorted in descending order, set the start to the last row of the sheet
        stop = 1                                                                                            #   If the data is to be sorted in descending order, set the stop to the first row of the sheet
    else:
        step = 1                                                                                            #   If the data is to be sorted in ascending order, set the step to 1
        start = 2                                                                                           #   If the data is to be sorted in ascending order, set the start to the first row of the sheet
        stop = sheet.nrows +1                                                                               #   If the data is to be sorted in ascending order, set the stop to the last row of the sheet
    for column_number in range(sheet.ncols):                                                                #   Loops through the columns in the XLSX file
        if array_header == "Date":
            relevant_column = column_number
            for row in range(start, stop, step):                                                            #   Loops through the rows in the XLSX file; This only executes if the Column Name provided returns a match
                requested_array.append(xlrd.xldate_as_datetime(sheet.cell_value(row - 1,relevant_column)).datetime.strftime("%d.%m.%y")) #   Appends the data from the XLSX file to the array
        elif sheet.cell_value(0,column_number) == array_header:                                             #   Using the Column Name provided, checks if the column is the one we want
            relevant_column = column_number                                                                 #   If it is, set the relevant_column to the column number
    if relevant_column == None:                                                                             #   If the Column Name provided returns no matches, we need to inform EndUser
        print("Error: Column not found")                                                                    #   Prints an error message to the console
        return None                                                                                         #   Returns None
    for row in range(start, stop, step):                                                                    #   Loops through the rows in the XLSX file; This only executes if the Column Name provided returns a match
        requested_array.append(sheet.cell_value(row - 1,relevant_column))                                   #   Appends the data from the XLSX file to the array
    return(requested_array)                                                                                 #   Returns the array
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for formatting the data in the DataFrame to be standardised for later parsing;
#   Utilizes pandas and numpy
#
#
def df_formatter(df):
    df["Lot/Inv_#"] = df["Lot/Inv_#"].fillna('').astype(str).str.replace(".0","",regex=False)               #   Replaces the .0 from the Lot/Inv_# column with an empty string; Our Lot/Inv_# numbers are supposed to be strings of integers, but the .0 is causing issues with the parsing and continuity of the data
    df["Vehicle"] = df["Vehicle"].mask(df["Vehicle"] == "", "RELIST")                                       #   Replaces the empty strings in the Vehicle column with the string "RELIST"; The RELIST status is denoted in the Vehicle Description column, but we wanted the Vehicle column for most of the data, so we'll fill in the blanks as we know what caused them.
    df = df.mask(df == "", "-")                                                                             #   Replaces the empty strings in the DataFrame with the string "-"; After the RELIST mask, we should only have empties in the CAT and Reg columns, so we'll replace them with the string "-" until later
    df = df.mask(df == np.nan, "-")                                                                         #   Replaces the NaN values in the DataFrame with the string "-"; After the RELIST mask, we should only have empties in the CAT and Reg columns, so we'll replace them with the string "-" until later
    df = df.reset_index(drop=True)                                                                          #   Resets the index of the DataFrame, so that it is in the correct order, just in case
    xlsx_injector(df)                                                                                       #   Calls the xlsx_injector function, which will inject the DataFrame into the XLSX file
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for preparing injection of the DataFrame into the 'COD_Vehicle_List.xlsx' file;
#   Utilizes pandas
#
#
def xlsx_injector(df):
    target_xlsx = pandas.read_excel(curr_dir+"\COD_Vehicle_List.xlsx", sheet_name="Sheet1")                 #   Opens the 'COD_Vehicle_List.xlsx' file
    beginning_row = target_xlsx.shape[1] +1                                                                 #   Sets the beginning row to the number of rows in the 'COD_Vehicle_List.xlsx' file + 1
    for row, column in target_xlsx.iterrows():                                                              #   Loops through the rows and columns in the 'COD_Vehicle_List.xlsx' file
        for item in df["Lot/Inv_#"].values:                                                                 #   Loops through the Lot/Inv_# values in the DataFrame
            if str(item) in str(column["Lot/Inv_#"]):                                                       #   Looks for a Lot/Inv_# overlap between the DataFrame and the 'COD_Vehicle_List.xlsx' file
                df = df.drop(df[df["Lot/Inv_#"].values == item].index)                                      #   Drops the row from the DataFrame that has the Lot/Inv_# overlap
    df = df.reset_index(drop = True)                                                                        #   Resets the index of the DataFrame, so that it is in the correct order, just in case
    print(df)                                                                                               #   Prints the DataFrame to the console, this is the penultimate cutdown of the DataFrame, before it is injected into the XLSX file
    Invoice_fetcher(df, beginning_row)                                                                      #   Calls the Invoice_fetcher function; Time to grab those last bits of data!
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------   
#
#                                               Penultimate Module: Invoice Fetcher
#
#           Purpose: Fetches the Invoice from the Copart website and saves it to the 'COD_Vehicle_List.xlsx' file, this includes Reg and CAT data
#
#           Elevated Caller:
#               Invoice_fetcher()
#                               -Parameters: df, beginning_row
#
#           Functions:
#               Invoice_extractor()
#                       - Extracts data from the Invoice
#                               -Parameters: file_location
#               cat_extractor()
#                       - Extracts the CAT data from the Invoice
#                               -Parameters: file_location
#
#           Order of Execution:
#               Invoice_fetcher()
#               |   Invoice_extractor()
#               |   cat_extractor()
#
#           Connection Call:
#               Invoice_fetcher()
#                   - Calls the append_df_to_excel() and Incomplete_Lot/Inv_#_handler() functions, located in the 'Data to Excel' Module
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for fetching the VAT Invoice from the Copart website, using 'new' Lot/Inv_# Numbers as the search index;
#   Utilizes pandas, os, time, shutil, Exception, selenium( WebDriverWait, Expected Conditions, By, webdriver )
#
#
def Invoice_fetcher(df, beginning_row):
    Reg_list = []                                                                                           #   Creates an empty array for the Registration Numbers
    LLT_list = []                                                                                           #   Creates an empty array for the Left Location Times
    cat_list = []                                                                                           #   Creates an empty list to store the CAT data; Similarly to the above
    for id in df["Lot/Inv_#"]:                                                                              #   The best way to look for a Vehicle overlap is to sift through the shorter list per car in the Large Data Set
        row = df[df["Lot/Inv_#"] == id].index                                                               #   Grabbing the position of the current car in the DataFrame that will hopefully be appended to the Excel file
        print(row)                                                                                          #   Provides the EndUser information on the current car being processed, this helps them track progress
        loop = 0                                                                                            #   Setting up a loop variable so that we can attempt to fetch the Invoice multiple times if it fails
        while loop <= 3:                                                                                    #   Try up to 3 times to fetch the Invoice, if it fails, then we will move on to the next car
            if loop<=2:                                                                                     #   If the loop is less than 2, then we will attempt to fetch the Invoice again; This will result in 3 attempts
                loop = loop + 1                                                                             #   Incrementing the loop variable by 1, this will allow us to try again if the Invoice fails to fetch
                call = id                                                                                   #   Originally had to manipulate the id variable to make it a string, but this is not necessary anymore; Lazy Dev (;-;)
                fetched = False                                                                             #   Setting up a boolean variable to check if the Invoice was successfully fetched
                try:                                                                                        #   If the VAT Invoice already exists for this Vehicle in the 'Payments_History' folder; Pull data from it
                    if df.iloc[df.index.get_loc(df.index[df["Lot/Inv_#"] == id][0])]["Vehicle"] != "RELIST":#   If the Vehicle is not a RELIST, we will search for an invoice
                        print("Vehicle Detected: Attempt " + str(int(loop)))                                #   Provides the EndUser information on the current car being processed, this helps them track progress
                        if os.path.exists(curr_dir+"\Payments_History\Invoice_"+id+ ".pdf"):                #   If the Invoice already exists, then we will pull the data from it; A reminder that existence of the Invoice
                            print("VAT Invoice already exists; Processing the PDF")                         #   Updates the EndUser on what path the script has taken
                            file_location = curr_dir+"\Payments_History\Invoice_"+id+ ".pdf"                #   Sets the file location to the Invoice
                            Reg = Invoice_extractor(file_location)                                          #   Sends the location to the Invoice Reg extracter function
                            Reg_list.append(Reg)                                                            #   Appending the recieved Reg data to the Reg list
                            Category = cat_extractor(file_location)                                         #   Sends the location to the CAT extracter function
                            cat_list.append(Category)                                                       #   Appending the recieved CAT data to the CAT list
                            LLT = LLT_fetcher(file_location)                                                #   Sending the DataFrame and the file location to the LLT fetcher function
                            LLT_list.append(LLT)
                            print(LLT)
                            print("Lot/Inv_# " + id + " has been processed")                                      #   Updates the EndUser on completed processing
                            loop = 99                                                                       #   Setting the loop variable to 99 to break out of the loop
                            continue                                                                        #   Continuing to the next iteration of the loop; Go to the next car in the DataFrame
                    
                        try:                                                                                #   If the Invoice does not exist, then we will attempt to fetch it; First via our 'Payment History' hub page
                            nav_to_payment_page(web)                                                        #   Navigates to the Payment History page
                            WebDriverWait(web, 5).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), '%s')]" % call)))
                            print("Lot/Inv_# Number found on page")                                               #   After going to the Payment History page, we will search for the Vehicle Lot/Inv_# Number, if found then we will click on it as this will take us to the VAT Invoice (We dont want a normal Invoice)
                            web.execute_script("arguments[0].scrollIntoView();", web.find_element(By.XPATH, "//*[contains(text(), '%s')]" % call))
                            web.find_element(By.XPATH, "//*[contains(text(), '%s')]" % call).click()        #   Clicking on the Vehicle Lot/Inv_# Number
                            print("Lot/Inv_# number clicked")                                                     #   Letting the EndUser know that the Lot/Inv_# Number was clicked; Helped me debug too
                            time.sleep(2)                                                                   #   Waiting 2 seconds for the Invoice to load; sweet peace and quiet
                            if web.current_url == "https://www.copart.co.uk/Lot/Inv_#/" +id:                      #   When a Lot/Inv_# is New/Incomplete a VAT Invoice will not be available, if this is the case, the Lot/Inv_# Number will redirect to Lot/Inv_# information.
                                print("Invoice not yet available")                                          #   Letting the EndUser know that the Invoice is not yet available
                                df = df.drop(df[df["Lot/Inv_#"] == id].index)                               #   As the Lot/Inv_# is too new to process, we will drop it from the DataFrame; This means that in a future run of the script, the Lot/Inv_# will be added when it's 'ripe' so to speak
                                loop = 999                                                                  #   Setting the loop variable to 999 to break out of the loop; No point in trying to fetch the Invoice again
                                continue                                                                    #   Continuing to the next iteration of the loop; Go to the next car in the DataFrame
                            WebDriverWait(web, 5).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Print')]")))
                            time.sleep(7)                                                                   #   If however a VAT Invoice is available, we will wait 7 seconds for it to load, after the Print button is detected; Since the script is running out of hours, time isn't a great issue. Hopefully this reduces the burden on the server.
                            web.execute_script(""" 
                                if (typeof window.print === 'undefined') {    
                                    setTimeout(function(){window.print();}, 1000);
                                } else {
                                    window.print();
                                }
                                """)
                            time.sleep(3)                                                                   #   The above Javascript code will 'print' the Invoice, we will wait 3 seconds for it to load; Thankfully our chrome_options allows us to default to 'Save as PDF', so we have no need to meddle with the settings in the print window.
                            if os.path.exists(curr_dir+"\Payments_History") == False:                       #   If the Payments_History folder does not exist...
                                os.mkdir(curr_dir+"\Payments_History")                                      #   Create the Payments_History folder
                            shutil.move(downloads_dir+"\Salvage Car Auctions & Used Vehicles _ Buy Online _ Copart UK.pdf", curr_dir+"\Payments_History\Invoice_"+id+ ".pdf")
                            file_location = curr_dir+"\Payments_History\Invoice_"+id+ ".pdf"                #   Sets the file location to the Invoice
                            print("VAT Invoice Downloaded via a 'Payment History' webpage Hyperlink; Processing the PDF")
                            Reg = Invoice_extractor(file_location)                                          #   Sends the location to the Invoice Reg extracter function
                            print(Reg)                                                                      #   Printing the Reg data to the EndUser
                            Reg_list.append(Reg)                                                            #   Appending the recieved Reg data to the Reg list
                            Category = cat_extractor(file_location)                                         #   Sends the location to the CAT extracter function
                            print(Category)                                                                 #   Printing the CAT data to the EndUser
                            cat_list.append(Category)                                                       #   Appending the recieved CAT data to the CAT list
                            LLT = LLT_fetcher( file_location)                                               #   Sending the DataFrame and the file location to the LLT fetcher function
                            LLT_list.append(LLT)
                            print(LLT)
                            print("Lot/Inv_# " + id + " has been processed")                                #   Updates the EndUser on completed processing
                            loop = 99                                                                       #   Setting the loop variable to 99 to break out of the loop
                            pass                                                                            #   Passing the loop; Go to the next car in the DataFrame  
                        except Exception as e:                                                              #   If the Invoice is STILL not available, then we will attempt to fetch it via a direct URL
                            
                            try:                                                                            #   Trying to fetch the Invoice via a direct URL; If Exception is raised, the script will not die
                                print("Unable to fetch VAT Invoice via a 'Payment History' webpage Hyperlink; The program will now attempt to use a 'Backdoor URL'")
                                web.get("https://www.copart.co.uk/member-payments/view-invoice/"+id+FSP_XID.post_id_hyperlink)
                                if "https://www.copart.co.uk/member-payments/view-invoice/" in web.current_url:
                                    WebDriverWait(web, 5).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Print')]")))
                                    time.sleep(2)                                                           #   Waiting 2 seconds for the Invoice to load; sweet peace and quiet
                                    web.execute_script("""
                                    if (typeof window.print === 'undefined') {    
                                        setTimeout(function(){window.print();}, 1000);
                                    } else {
                                        window.print();
                                    }
                                    
                                    """)
                                    time.sleep(3)                                                           #   The above Javascript code will 'print' the Invoice, we will wait 3 seconds for it to load; Thankfully our chrome_options allows us to default to 'Save as PDF', so we have no need to meddle with the settings in the print window.
                                    if os.path.exists(curr_dir+"\Payments_History") == False:               #   If the Payments_History folder does not exist...
                                        os.mkdir(curr_dir+"\Payments_History")                              #   Create the Payments_History folder
                                    if os.path.exists(downloads_dir+"\Online Vehicle Auctions Copart UK_ Salvage & Used Cars, Trucks & More.pdf"):
                                        shutil.move(downloads_dir+"\Online Vehicle Auctions Copart UK_ Salvage & Used Cars, Trucks & More.pdf", curr_dir+"\Payments_History\Invoice_"+id+ ".pdf")
                                    elif os.path.exists(downloads_dir+"\Salvage Car Auctions & Used Vehicles _ Buy Online _ Copart UK.pdf"):
                                        shutil.move(downloads_dir+"\Salvage Car Auctions & Used Vehicles _ Buy Online _ Copart UK.pdf", curr_dir+"\Payments_History\Invoice_"+id+ ".pdf")
                                    else:
                                        pass                                                                #   If an Invoice is downloaded, it will be moved to the Payments_History folder; If not, the script will pass
                                    print("VAT Invoice Downloaded via a 'Backdoor URL'; Processing the PDF")
                                    file_location = curr_dir+"\Payments_History\Invoice_"+id+ ".pdf"        #   Sets the file location to the Invoice
                                    Reg = Invoice_extractor(file_location)                                  #   Sends the location to the Invoice Reg extracter function
                                    print(Reg)                                                              #   Printing the Reg data to the EndUser
                                    Reg_list.append(Reg)                                                    #   Appending the recieved Reg data to the Reg list
                                    Category = cat_extractor(file_location)                                 #   Sends the location to the CAT extracter function
                                    print(Category)                                                         #   Printing the CAT data to the EndUser
                                    cat_list.append(Category)                                               #   Appending the recieved CAT data to the CAT list
                                    LLT = LLT_fetcher(file_location)                                        #   Sending the DataFrame and the file location to the LLT fetcher function
                                    LLT_list.append(LLT)
                                    print(LLT)
                                    print("Lot/Inv_# " + id + " has been processed")                        #   Updates the EndUser on completed processing
                                    loop = 99                                                               #   Sets loop variable to 99 to break out of the loop
                                    pass                                                                    #   Passing the loop; Go to the next car in the DataFrame
                                else:
                                    print("Backdoor URL Access was denied by Copart")                       #   If the Backdoor URL Access was denied by Copart, the script will print this message; It'll proceed to next loop if loop isnt already at max
                                    
                                    pass
                            except Exception as e:                                                          #   If the Invoice is STILL not available, I have no idea what went wrong...; Script will print the Exception and proceed to next loop if loop isnt already at max
                                print(e)
                                pass
                    else:                                                                                   #   Now that the Vehicles are handled, lets handle the RELIST entries
                        print('Item of type: "RELIST" was detected, skipping search for VAT Invoice and CAT on this item')
                        cat_list.append("-")                                                                #   Filling NaNs with -
                        print("Lot/Inv_# " + id + " has been processed")                                          #   Updates the EndUser on completed processing
                        loop = 99                                                                           #   Sets loop variable to 99 to break out of the loop
                        pass                                                                                #   Passing the loop; Go to the next car in the DataFrame
                except Exception:                                                                           #   If the Item is a relist, but has caused a TimeoutException, the script will print this message; It'll proceed to next loop if loop isnt already at max
                    if df.iloc[df.index.get_loc(df.index[df["Lot/Inv_#"] == id][0])]["Vehicle"] == "RELIST":
                        print("TimeoutException caused by a RELIST")                                        #   If the TimeoutException was caused by a RELIST, the script will print this message; It'll proceed to next loop if loop isnt already at max
                        cat_list.append("-")                                                                #   Filling NaNs with -; just in case
                        print("Lot/Inv_# " + id + " has been processed")                                          #   Updates the EndUser on completed processing
                        loop = 99                                                                           #   Sets loop variable to 99 to break out of the loop
                        pass                                                                                #   Passing the loop; Go to the next car in the DataFrame
                    elif Exception.__name__ == "TimeoutException" and df.iloc[df.index.get_loc(df.index[df["Lot/Inv_#"] == id][0])]["Vehicle"] != "RELIST":
                        print("""A TimeoutException has occured for an item other than a relist, this will be fixed in another run of the loop""")
                        
            else:                                                                                           #   If the loop is not at max, the script will print this message; Next Vehicle please!
                print("The current item has been tried to be added 3 times, it seems that this item's Reg is not available to the program")
                df = df.drop(df[df["Lot/Inv_#"] == id].index)                                               #   Dropping the Lot/Inv_# as it's probably not ready for processing, which is why we are unable to process it; This will prevent the script from trying to process the same Lot/Inv_# multiple times or not at all
                break                                                                                       #   Breaking out of the loop
    print(Reg_list)                                                                                         #   Printing the Reg list to the EndUser
    print(cat_list)                                                                                         #   Printing the CAT list to the EndUser
    print(LLT_list)                                                                                         #   Printing the LLT list to the EndUser
    df["Registration"] = pandas.Series(Reg_list)                                                            #   Adding the Reg list to the DataFrame
    df["Category"] = pandas.Series(cat_list)                                                                #   Adding the CAT list to the DataFrame
    df["LLT"] = pandas.Series(LLT_list)                                                                     #   Adding the LLT list to the DataFrame
    df.replace(to_replace=[None], value=np.nan, inplace=True)                                               #   Replacing the None values with NaNs
    #df.dropna(axis=0, how='any', inplace=True)                                                             #   Dropping the Incomplete Lots from the DataFrame; We don't want those in our 'COD_Vehicle_List' DataFrame
    print(df)                                                                                               #   Printing the DataFrame to the EndUser one last time; This is what they should expect to be added to the bottom of the file
    if df.empty == False:                                                                                   #   If the DataFrame is not empty, the script will print this message
        print("Reg and CAT values have been appended")                                                      #   Reg and CAT were successful; If they were not the Lots would've been flagged as Incomplete and dropped from the DataFrame, leaving an empty DataFrame
    append_df_to_excel("COD_Vehicle_List.xlsx", df, sheet_name = "Sheet1", startrow = beginning_row) #   Calling for the append_df_to_excel function to append the DataFrame to the Excel file
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for extracting the Reg data from the Invoice;
#   Utilizes os and PyPDF2
#
#       
def Invoice_extractor(file_location):
    if os.path.exists(file_location):                                                                       #   If the file exists, we can proceed
        pdf_text = PyPDF2.PdfFileReader(file_location)                                                      #   Use PyPDF2 to extract the text from the PDF
        num_pages = pdf_text.getNumPages()                                                                  #   Get the number of pages in the PDF
        for page in range(0, num_pages):                                                                    #   Loop through each page in the PDF   
            page_content = pdf_text.getPage(page).extractText()                                             #   Extract the text from the PDF
            if "VRN:" in page_content:                                                                      #   If the text contains the word "VRN:", we can proceed to fetch what comes after
                Content_after_Reg = page_content.split("VRN:")[1]                                           #   Split the text after the word "VRN", our Reg Total will be in the second part of the split, so we take [1] instead of [0]
                Unprocesed_Reg_string = Content_after_Reg.split("\n")[0]                                    #   Split again on a new line, the VRN will be in the first part of the split, so we take [0] instead of [1]
                Split_space = Unprocesed_Reg_string.split(" ")                                              #   Split the string on spaces, giving us an array of spaces and our Reg Total
                print(Split_space[2])                                                                       #   Print the Reg to the EndUser
                return(Split_space[2])                                                                      #   Return the Reg to the function
            else:                                                                                           #   If the page does not contain the word "Reg Total:"
                print("No Reg on page " + str(page))                                                        #   Print the page number to the EndUser, the script will continue to the next page
    else:                                                                                                   #   If the file does not exist, we can't proceed
        Reg = ""                                                                                            #   Set the Reg variable to an empty string
        return(Reg)                                                                                         #   Return the Reg Total
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
#   Function for extracting the CAT data from the Invoice;
#   Utilizes os and PyPDF2
#
#   
def cat_extractor(file_location):
    if os.path.exists(file_location):                                                                       #   If the file exists, we can proceed
        pdf_text = PyPDF2.PdfFileReader(file_location)                                                      #   Use PyPDF2 to extract the text from the PDF
        num_pages = pdf_text.getNumPages()                                                                  #   Get the number of pages in the PDF
        for page in range(0, num_pages):                                                                    #   Loop through each page in the PDF
            page_content = pdf_text.getPage(page).extractText()                                             #   Extract the text from the PDF
            if "ABI Category: " in page_content:                                                            #   If the text contains the word "ABI Category: ", we can proceed to fetch what comes after
                Content_after_category = page_content.split("ABI Category: ")[1]                            #   Split the text after the word "ABI Category: ", our ABI Category will be in the second part of the split, so we take [1] instead of [0]
                Category = Content_after_category.split(" ")[0]                                             #   Split again on a new line, the ABI Category will be in the first part of the split, so we take [0] instead of [1]
                return(Category)                                                                            #   Return the ABI Category
            else:                                                                                           #   If the page does not contain the word "ABI Category:"
                print("No CAT on page " + str(page))                                                        #   Print the page number to the EndUser, the script will continue to the next page
    else:                                                                                                   #   If the file does not exist, we can't proceed
        Category = ""                                                                                       #   Set the Category variable to an empty string                                                                                       
        return(Category)                                                                                    #   Return the ABI Category
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#   Function to grab Colour
#
def colour_fetcher(df):
    color_list = []                                                                                         #   Creating an empty list for the colours
    for lot in df["Lot/Inv_#"]:                                                                             #   Loop through each Lot in the DataFrame       
        web.get("https://www.copart.co.uk/lot/" + lot)                                                      #   Open the lot in the browser
        WebDriverWait(web, 100).until(EC.presence_of_element_located((By.XPATH, '//span[@data-uname="lotdetailColorvalue"]')))
        color = web.find_element(By.XPATH, '//span[@data-uname="lotdetailColorvalue"]').text                #   Find the element with the colour and get the innerHTML
        exploded_color = color.strip(" ")                                                                   #   If the element is not an empty string, a space or two spaces
        color_list.append(str(exploded_color))                                                              #   If the element is a letter, append it to the list
    return(color_list)
#
#
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#   Function to fetch LLT
#
def LLT_fetcher(file_location):
    if os.path.exists(file_location):                                                                       #   If the file exists, we can proceed
        pdf_text = PyPDF2.PdfFileReader(file_location)                                                      #   Use PyPDF2 to extract the text from the PDF
        num_pages = pdf_text.getNumPages()                                                                  #   Get the number of pages in the PDF
        for page in range(0, num_pages):                                                                    #   Loop through each page in the PDF
            page_content = pdf_text.getPage(page).extractText()                                             #   Extract the text from the PDF
            if "Date: " in page_content:                                                                    #   If the text contains the word "Date: ", we can proceed to fetch what comes after
                Content_after_Date = page_content.split("Date:")[1]                                         #   Split the text after the word "Date: ", our LLT will be in the second part of the split, so we take [1] instead of [0]
                Exploded_LLT = Content_after_Date.split(" ")                                                #   Split again on a new line, the LLT will be in the first part of the split, so we take [0] instead of [1]
                for i in Exploded_LLT:
                    if i != "":
                        LLT_unformatted = i
                        if len(i) == 8:
                            LLT = LLT_unformatted[0:2] + "." + LLT_unformatted[3:5] + "." + LLT_unformatted[6:8]
                        elif len(i) == 7:
                            LLT = "0" + LLT_unformatted[0:1] + "." + LLT_unformatted[2:4] + "." + LLT_unformatted[5:7]
                        return(LLT)                                                                         #   Return the LLT
            else:                                                                                           #   If the page does not contain the word "Date:"
                print("No LLT on page " + str(page))                                                        #   Print the page number to the EndUser, the script will continue to the next page
    else:                                                                                                   #   If the file does not exist, we can't proceed
        LLT = ""                                                                                            #   Set the Category variable to an empty string                                                                                       
        return(LLT)                                                                                         #   Return the LLT


    
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#
#
def append_df_to_excel(filename, df, sheet_name , startrow):
    color_list = colour_fetcher(df)
    df = df.assign(Colour = color_list)
    print(df)
    workbook = openpyxl.load_workbook(curr_dir+"\COD_Vehicle_List.xlsx")                                    #   Load the Excel file into memory
    worksheet = workbook[sheet_name]                                                                        #   Select the sheet we want to append to
    for row in df.itertuples():                                                                             #   Loop through the DataFrame
        worksheet.append(row[1:])                                                                           #   Append the row (except for the index) to the Excel file
    workbook.save("COD_Vehicle_List.xlsx")                                                                      #   Save the Excel file
    print("done")


copart_90_day_csv_download()